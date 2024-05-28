import sys
import openpyxl

PATH = '/home/lurr3t/exjobb/src/logs/output.xlsx'


class ExcelParser:
    def __init__(self):
        # load workbook
        self.wb = openpyxl.load_workbook(PATH)
        # call the active worksheet
        self.ws = self.wb.active

        # clear sheet
        self.ws.delete_cols(1, 50)
        self.ws.delete_rows(1, 200)

        # create first row with names
        new_row = ("Size", "quic", "quic_compress", "tcp", "tcp_compress")
        self.ws.append(new_row)
        self.wb.save(PATH)




    # episode = [["quic", 0, 0, []], ["quic_compress", 0, 0, []], ["tcp", 0, 0, []], ["tcp_compress", 0, 0, []], 0]
    def load_episode(self, episode):

        # load cell with data
        new_row = (episode[4], episode[0][2], episode[1][2],
                   episode[2][2], episode[3][2])
        self.ws.append(new_row)

        #ws['K1'] = 'Kb'
        #ws.cell(row=1, column=12, value='Sum of Sales')

        self.wb.save(PATH)









